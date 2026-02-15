"""
ARGOS - Herramientas Excel
Lectura y escritura de archivos Excel con openpyxl.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl


def read_excel_summary(path):
    """Lee un Excel y retorna resumen de hojas con dimensiones."""
    wb = openpyxl.load_workbook(path, data_only=True)
    summary = {}
    for name in wb.sheetnames:
        ws = wb[name]
        summary[name] = {
            'rows': ws.max_row,
            'cols': ws.max_column,
        }
    wb.close()
    return summary


def read_sheet_data(path, sheet_name, row_start=1, row_end=None,
                    col_start=1, col_end=None, data_only=True):
    """
    Lee datos de una hoja específica.

    Returns:
        lista de listas con los valores de cada celda
    """
    wb = openpyxl.load_workbook(path, data_only=data_only)
    ws = wb[sheet_name]

    if row_end is None:
        row_end = ws.max_row
    if col_end is None:
        col_end = ws.max_column

    data = []
    for row in range(row_start, row_end + 1):
        row_data = []
        for col in range(col_start, col_end + 1):
            row_data.append(ws.cell(row=row, column=col).value)
        data.append(row_data)

    wb.close()
    return data


def update_cells(path, sheet_name, changes):
    """
    Actualiza celdas específicas de un Excel.

    Args:
        path: ruta al archivo Excel
        sheet_name: nombre de la hoja
        changes: lista de dicts con keys 'row', 'col', 'value'
    """
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]

    for change in changes:
        ws.cell(row=change['row'], column=change['col']).value = change['value']
        print(f"  R{change['row']}C{change['col']} = {change['value']}")

    wb.save(path)
    print(f'Guardado: {path}')
